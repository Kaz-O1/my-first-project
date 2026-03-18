#!/usr/bin/env python3
"""
Invoice Processor – Ohad Kazoom (עוסק מורשה) & Cril-Tech
---------------------------------------------------------
Hybrid OCR pipeline: pytesseract first, Claude API fallback.
Supports two entities (ohad / cril-tech), each with its own
incoming folder, registry, processed tree and monthly Excel reports.
Generates a single HTML dashboard covering all entities.
"""

import argparse
import base64
import hashlib
import io
import json
import logging
import re
import shutil
import subprocess
import webbrowser
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import anthropic

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Paths & constants ─────────────────────────────────────────────────────────

BASE_DIR      = Path(__file__).parent
DASHBOARD_PATH = BASE_DIR / "processed" / "dashboard.html"

ENTITIES: dict[str, dict] = {
    "ohad": {
        "label":          "Ohad Kazoom",
        "incoming_dir":   BASE_DIR / "incoming_invoices" / "ohad",
        "processed_base": BASE_DIR / "processed" / "ohad",
        "registry_path":  BASE_DIR / "processed_ohad.json",
    },
    "cril-tech": {
        "label":          "Cril-Tech",
        "incoming_dir":   BASE_DIR / "incoming_invoices" / "cril-tech",
        "processed_base": BASE_DIR / "processed" / "cril-tech",
        "registry_path":  BASE_DIR / "processed_cril.json",
    },
}

SUPPORTED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"}

CATEGORIES = [
    "fuel", "entertainment", "office supplies", "vehicle",
    "communication", "professional services", "other",
]

MONTH_NAMES = {
    1: "01-January",   2: "02-February",  3: "03-March",
    4: "04-April",     5: "05-May",       6: "06-June",
    7: "07-July",      8: "08-August",    9: "09-September",
    10: "10-October",  11: "11-November", 12: "12-December",
}

OCR_CONFIDENCE_THRESHOLD = 92.0  # minimum average word confidence %


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class InvoiceData:
    file_name:        str
    file_hash:        str
    invoice_date:     Optional[str]   = None   # ISO: YYYY-MM-DD
    vendor_name:      Optional[str]   = None
    invoice_number:   Optional[str]   = None
    total_amount:     Optional[float] = None   # NIS incl. VAT (after conversion)
    vat_amount:       Optional[float] = None   # מע"מ in NIS (after conversion)
    currency:         str             = "ILS"  # original invoice currency
    original_amount:  Optional[float] = None   # amount in original currency (pre-conversion)
    exchange_rate:    Optional[float] = None   # ILS per 1 unit of foreign currency
    category:         str             = "other"
    ocr_method:       str             = "unknown"
    processed_at:     str             = ""
    destination_path: str             = ""

    def __post_init__(self):
        if not self.processed_at:
            self.processed_at = datetime.now().isoformat()


# ── File utilities ────────────────────────────────────────────────────────────

def file_sha256(file_path: Path) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def open_as_pil_images(file_path: Path) -> list:
    from PIL import Image
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        try:
            from pdf2image import convert_from_path
            return convert_from_path(str(file_path), dpi=200)
        except Exception as e:
            logger.warning(f"pdf2image failed for {file_path.name}: {e}")
            return []
    else:
        try:
            img = Image.open(file_path)
            img.load()
            return [img]
        except Exception as e:
            logger.warning(f"Cannot open {file_path.name}: {e}")
            return []


def to_png_base64(file_path: Path) -> tuple[str, str]:
    suffix = file_path.suffix.lower()
    NATIVE_TYPES = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".png": "image/png",  ".gif":  "image/gif", ".webp": "image/webp",
    }
    if suffix == ".pdf":
        images = open_as_pil_images(file_path)
        if images:
            buf = io.BytesIO()
            images[0].save(buf, format="PNG")
            return base64.standard_b64encode(buf.getvalue()).decode(), "image/png"
        with open(file_path, "rb") as f:
            return base64.standard_b64encode(f.read()).decode(), "application/pdf"
    # For all image types: open with PIL to detect the *actual* format
    # (avoids sending wrong media_type for mislabelled files, e.g. .png that is JPEG)
    images = open_as_pil_images(file_path)
    if images:
        img = images[0]
        actual_fmt = img.format or "PNG"   # PIL detects real format
        SAVE_MAP = {"JPEG": ("JPEG", "image/jpeg"), "WEBP": ("WEBP", "image/webp")}
        save_fmt, media = SAVE_MAP.get(actual_fmt, ("PNG", "image/png"))
        buf = io.BytesIO()
        img.save(buf, format=save_fmt)
        return base64.standard_b64encode(buf.getvalue()).decode(), media
    raise ValueError(f"Cannot encode {file_path.name} for Claude API")


# ── Tesseract OCR ─────────────────────────────────────────────────────────────

def tesseract_ocr(file_path: Path) -> tuple[str, float]:
    import pytesseract
    images = open_as_pil_images(file_path)
    if not images:
        return "", 0.0
    all_text, total_conf, word_count = [], 0.0, 0
    for img in images:
        data = pytesseract.image_to_data(
            img, lang="heb+eng", output_type=pytesseract.Output.DICT
        )
        for conf in data["conf"]:
            if isinstance(conf, (int, float)) and conf > 0:
                total_conf += float(conf)
                word_count += 1
        all_text.append(pytesseract.image_to_string(img, lang="heb+eng"))
    avg_conf = (total_conf / word_count) if word_count > 0 else 0.0
    return "\n".join(all_text).strip(), avg_conf


# ── OCR validation patterns ───────────────────────────────────────────────────

_DATE_RE = re.compile(
    r"""
    \d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}
    | \d{4}[./\-]\d{1,2}[./\-]\d{1,2}
    | \d{1,2}\s+(?:ינואר|פברואר|מרץ|אפריל|מאי|יוני|
                  יולי|אוגוסט|ספטמבר|אוקטובר|נובמבר|דצמבר)\s+\d{4}
    | \d{1,2}\s+(?:January|February|March|April|May|June|July|
                  August|September|October|November|December)\s+\d{4}
    """,
    re.VERBOSE | re.IGNORECASE,
)

_AMOUNT_RE = re.compile(
    r"""
    ₪\s*[\d,]+(?:\.\d{1,2})?
    | [\d,]+(?:\.\d{1,2})?\s*₪
    | [\d,]+(?:\.\d{1,2})?\s*(?:ש[\"׳]ח|שח|NIS)
    | (?:סה[\"׳]כ|total)\s*:?\s*[\d,]+
    """,
    re.VERBOSE | re.IGNORECASE,
)

_INV_NUM_RE = re.compile(
    r"""
    (?:invoice|inv|חשבונית|מס[\'\"׳]?\s*חשבונית|מספר\s*חשבונית)\s*[:#]?\s*([A-Za-z0-9\-/]+)
    | (?:no\.?|number|מספר|מס\')\s*[:#]?\s*([A-Za-z0-9\-/]+)
    | \#\s*([A-Za-z0-9\-/]{4,})
    """,
    re.VERBOSE | re.IGNORECASE,
)


def validate_ocr(text: str, confidence: float) -> bool:
    if confidence < OCR_CONFIDENCE_THRESHOLD:
        logger.debug(f"OCR confidence too low: {confidence:.1f}%")
        return False
    has_date   = bool(_DATE_RE.search(text))
    has_amount = bool(_AMOUNT_RE.search(text))
    return has_date and has_amount


# ── Text-based field extraction ───────────────────────────────────────────────

def _parse_date(text: str) -> Optional[str]:
    from dateutil import parser as dp
    m = _DATE_RE.search(text)
    if m:
        try:
            return dp.parse(m.group(0), dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def _parse_total(text: str) -> Optional[float]:
    amounts = []
    for pattern in [
        r'(?:סה[\"׳]כ|סהכ|total\s+amount|amount\s+due|לתשלום)\s*:?\s*₪?\s*([\d,]+(?:\.\d{1,2})?)',
        r'₪\s*([\d,]+(?:\.\d{1,2})?)',
        r'([\d,]+(?:\.\d{1,2})?)\s*₪',
        r'([\d,]+(?:\.\d{1,2})?)\s*(?:ש[\"׳]ח|NIS)',
    ]:
        for m in re.finditer(pattern, text, re.IGNORECASE):
            try:
                amounts.append(float(m.group(1).replace(",", "")))
            except (ValueError, IndexError):
                pass
    return max(amounts) if amounts else None


def _parse_vat(text: str) -> Optional[float]:
    for pattern in [
        r'(?:מע[\"׳]מ|vat|tax)\s*:?\s*₪?\s*([\d,]+(?:\.\d{1,2})?)',
        r'([\d,]+(?:\.\d{1,2})?)\s*(?:מע[\"׳]מ)',
    ]:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(",", ""))
            except ValueError:
                pass
    return None


def _parse_invoice_number(text: str) -> Optional[str]:
    m = _INV_NUM_RE.search(text)
    if m:
        matched = next((g for g in m.groups() if g is not None), None)
        return matched.strip() if matched else None
    m = re.search(r'\b(\d{5,})\b', text)
    return m.group(1) if m else None


_CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "fuel":                 ["דלק", "fuel", "petrol", "gasoline", "תדלוק",
                             "paz", "dor", "sonol", "delek", "ten", "אמישראגז"],
    "entertainment":        ["מסעדה", "restaurant", "cafe", "קפה", "בילוי",
                             "entertainment", "אירוח", "bar", "pub", "pizza", "burger"],
    "office supplies":      ["משרד", "office", "paper", "נייר", "מדפסת", "printer",
                             "סטציונרי", "stationery", "toner", "ink", "cartridge",
                             "kravitz", "סטימצקי"],
    "vehicle":              ["רכב", "vehicle", "car", "garage", "מוסך", "טסט",
                             "ביטוח רכב", "parking", "חניה", "tyre", "tire", "גלגל", "שמן"],
    "communication":        ["טלפון", "phone", "internet", "סלולר", "cellular",
                             "bezeq", "בזק", "hot", "partner", "cellcom", "012", "019",
                             "yes ", "איקס", "xfone"],
    "professional services":["עו\"ד", "עורך דין", "רואה חשבון", "accountant",
                             "lawyer", "ייעוץ", "consulting", "notary", "נוטריון",
                             "professional", "architect", "אדריכל"],
}


def _categorize(text: str) -> str:
    text_lower = text.lower()
    for category, keywords in _CATEGORY_KEYWORDS.items():
        if any(kw.lower() in text_lower for kw in keywords):
            return category
    return "other"


def extract_from_text(text: str) -> dict:
    return {
        "invoice_date":   _parse_date(text),
        "vendor_name":    None,
        "invoice_number": _parse_invoice_number(text),
        "total_amount":   _parse_total(text),
        "vat_amount":     _parse_vat(text),
        "category":       _categorize(text),
    }


# ── Claude API extractor ──────────────────────────────────────────────────────

_SYSTEM = """\
You are an invoice data extraction specialist. You extract structured data \
from invoices in Hebrew and English. Respond ONLY with a valid JSON object, \
no markdown fences, no additional text.\
"""

_PROMPT = """\
Extract the following fields from this invoice image.

Return exactly this JSON structure:
{
  "invoice_date":   "YYYY-MM-DD or null",
  "vendor_name":    "company or person name, or null",
  "invoice_number": "invoice number string, or null",
  "currency":       "ISO currency code: ILS / USD / EUR / GBP / etc.",
  "total_amount":   <grand total in the invoice's OWN currency including VAT, or null>,
  "vat_amount":     <VAT amount in the invoice's OWN currency; 0 if no VAT; null if unknown>,
  "category":       "<one of the values below>"
}

Category values:
  fuel               – gas stations, fuel (דלק, תחנת דלק)
  entertainment      – restaurants, cafes, events (מסעדה, בילוי, קפה)
  office supplies    – stationery, printer supplies, paper (ציוד משרדי)
  vehicle            – car repairs, parking, car insurance (מוסך, חניה, רכב)
  communication      – phone, internet, cellular bills (טלפון, אינטרנט, סלולר)
  professional services – lawyers, accountants, consultants (עו\"ד, רו\"ח, ייעוץ)
  other              – anything else

Rules:
- Detect the currency from symbols ($, €, £, ₪) or text (USD, EUR, NIS, שקל).
- total_amount must be the grand total INCLUDING VAT (סה\"כ לתשלום).
- Amounts are plain numbers in the invoice's own currency (no symbol).
- If the invoice explicitly shows no VAT (foreign vendor, export invoice), set vat_amount to 0.
- If this appears to be a credit card charge screenshot (no formal invoice, just a total \
charge), the total already includes 18% Israeli VAT: set vat_amount = round(total*18/118, 2).
- Return ONLY the JSON object.\
"""


def extract_with_claude(file_path: Path, client: anthropic.Anthropic) -> dict:
    logger.info(f"Claude API: extracting from {file_path.name}")
    b64, media_type = to_png_base64(file_path)

    if media_type == "application/pdf":
        content = [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
            {"type": "text", "text": _PROMPT},
        ]
    else:
        content = [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
            {"type": "text", "text": _PROMPT},
        ]

    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        system=_SYSTEM,
        messages=[{"role": "user", "content": content}],
    ) as stream:
        response = stream.get_final_message()

    raw = next((b.text for b in response.content if b.type == "text"), "")
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    raw = re.sub(r"\s*```$",          "", raw.strip(), flags=re.MULTILINE)

    # If Claude wrapped the JSON in explanatory text, extract the {...} object
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        start, end = raw.find("{"), raw.rfind("}")
        if start != -1 and end != -1:
            data = json.loads(raw[start:end + 1])
        else:
            logger.debug(f"Claude raw response: {raw[:300]}")
            raise

    for key in ("total_amount", "vat_amount"):
        v = data.get(key)
        if v is not None:
            try:
                data[key] = float(str(v).replace(",", ""))
            except (ValueError, TypeError):
                data[key] = None

    currency = (data.get("currency") or "ILS").upper().strip()
    data["currency"] = currency if len(currency) <= 5 else "ILS"

    if data.get("category") not in CATEGORIES:
        data["category"] = "other"

    return data


# ── Bank of Israel exchange rate ──────────────────────────────────────────────

def get_boi_rate(currency: str, date_str: Optional[str]) -> Optional[float]:
    """Fetch שער יציג (official exchange rate) from Bank of Israel SDMX API.
    Falls back to previous days to handle weekends / holidays.
    Returns ILS per 1 unit of the foreign currency, or None on failure.
    """
    import urllib.request
    from datetime import datetime as _dt, timedelta

    if not currency or currency == "ILS":
        return 1.0

    try:
        base = _dt.strptime(date_str, "%Y-%m-%d") if date_str else _dt.now()
    except (ValueError, TypeError):
        base = _dt.now()

    for delta in range(7):
        check = (base - timedelta(days=delta)).strftime("%Y-%m-%d")
        url = (
            f"https://edge.boi.gov.il/FusionEdgeServer/sdmx/v2/data/"
            f"dataflow/BOI.STATISTICS/EXR/1.0/RER_{currency}_ILS"
            f"?format=sdmx-json&startperiod={check}&endperiod={check}"
        )
        try:
            req = urllib.request.Request(url, headers={"Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                payload = json.loads(resp.read())
            series = payload["data"]["dataSets"][0]["series"]
            if series:
                obs = list(series.values())[0].get("observations", {})
                if obs:
                    rate = list(obs.values())[0][0]
                    if rate:
                        logger.info(f"  BOI שער יציג {currency}/ILS on {check}: {rate}")
                        return float(rate)
        except Exception as e:
            logger.debug(f"BOI rate lookup failed ({check}): {e}")

    logger.warning(f"Could not fetch BOI rate for {currency} on {date_str}")
    return None


# ── Registry helpers ──────────────────────────────────────────────────────────

def load_registry(registry_path: Path) -> dict:
    if registry_path.exists():
        with open(registry_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_registry(registry: dict, registry_path: Path) -> None:
    with open(registry_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2, default=str)


# ── File organiser ────────────────────────────────────────────────────────────

def month_folder_for(invoice_date: Optional[str], processed_base: Path) -> Path:
    if invoice_date:
        try:
            dt = datetime.strptime(invoice_date, "%Y-%m-%d")
            return processed_base / str(dt.year) / MONTH_NAMES[dt.month]
        except (ValueError, KeyError):
            pass
    now = datetime.now()
    return processed_base / str(now.year) / MONTH_NAMES[now.month]


def copy_to_processed(src: Path, invoice: InvoiceData, processed_base: Path) -> Path:
    dest_dir = month_folder_for(invoice.invoice_date, processed_base)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / src.name
    counter = 1
    while dest.exists():
        dest = dest_dir / f"{src.stem}_{counter}{src.suffix}"
        counter += 1
    shutil.copy2(src, dest)
    logger.info(f"  → {dest.relative_to(BASE_DIR)}")
    return dest


# ── Excel report ──────────────────────────────────────────────────────────────

_HEADERS    = ["File Name", "Date", "Vendor", "Invoice #",
               "Total (₪)", "VAT (₪)", "Net (₪)", "Category", "OCR Method"]
_COL_WIDTHS = [32, 12, 26, 16, 13, 13, 13, 22, 14]


def generate_excel_report(month_folder: Path, invoices: list[InvoiceData]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = month_folder.name.split("-", 1)[-1] if "-" in month_folder.name else month_folder.name

    hdr_fill  = PatternFill("solid", fgColor="1F3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", fgColor="EEF2F7")
    num_align = Alignment(horizontal="right",  vertical="center")
    str_align = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    sum_fill  = PatternFill("solid", fgColor="1F3A5F")
    sum_font  = Font(bold=True, color="FFFFFF", size=11)

    for col, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, hdr_align, bdr
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, inv in enumerate(invoices, start=2):
        net = round(inv.total_amount - inv.vat_amount, 2) \
              if inv.total_amount is not None and inv.vat_amount is not None else None
        row = [
            inv.file_name, inv.invoice_date or "Unknown", inv.vendor_name or "Unknown",
            inv.invoice_number or "N/A", inv.total_amount, inv.vat_amount, net,
            inv.category, inv.ocr_method,
        ]
        alt = row_idx % 2 == 0
        for col, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.border = bdr
            if alt:
                c.fill = alt_fill
            if col in (5, 6, 7):
                c.alignment = num_align
                if value is not None:
                    c.number_format = '#,##0.00'
            else:
                c.alignment = str_align

    last_data = len(invoices) + 1
    if last_data >= 2:
        total_row = last_data + 1
        for col in range(1, len(_HEADERS) + 1):
            c = ws.cell(row=total_row, column=col)
            c.fill, c.font, c.border = sum_fill, sum_font, bdr
            if col == 1:
                c.value, c.alignment = "TOTAL", str_align
            elif col in (5, 6, 7):
                cl = get_column_letter(col)
                c.value = f"=SUM({cl}2:{cl}{last_data})"
                c.number_format, c.alignment = '#,##0.00', num_align

    ws_meta = wb.create_sheet("Summary")
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 22
    meta_hdr = Font(bold=True, color="1F3A5F")
    for r, (label, value) in enumerate([
        ("Report Month",       month_folder.name),
        ("Report Generated",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Invoice Count",      len(invoices)),
        ("Total Expenses (₪)", sum(i.total_amount or 0 for i in invoices)),
        ("Total VAT (₪)",      sum(i.vat_amount   or 0 for i in invoices)),
    ], start=1):
        a = ws_meta.cell(row=r, column=1, value=label)
        a.font = meta_hdr
        ws_meta.cell(row=r, column=2, value=value)

    excel_path = month_folder / "monthly_report.xlsx"
    wb.save(excel_path)
    logger.info(f"Excel report → {excel_path.relative_to(BASE_DIR)}")


# ── HTML Dashboard ────────────────────────────────────────────────────────────

_CAT_COLORS = {
    "fuel":                  "#E74C3C",
    "entertainment":         "#E67E22",
    "office supplies":       "#3498DB",
    "vehicle":               "#2ECC71",
    "communication":         "#9B59B6",
    "professional services": "#1ABC9C",
    "other":                 "#95A5A6",
}


def _needs_review(data: dict) -> bool:
    """True if this registry entry should appear in the manual-review panel."""
    return (
        data.get("ocr_method") not in ("manual-skip",)
        and (
            data.get("ocr_method") in ("failed", "unknown")
            or data.get("total_amount") is None
            or data.get("invoice_date") is None
        )
    )


def _build_dashboard_data(all_registries: dict) -> dict:
    """Aggregate registry data into the structure the HTML dashboard expects."""
    result: dict = {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "entities":  {},
    }
    for entity_key, (entity_cfg, registry) in all_registries.items():
        months: dict = {}
        review: list = []

        for fhash, data in registry.items():
            # Always collect incomplete / failed invoices for the review panel
            if _needs_review(data):
                review.append({
                    "file_hash":      fhash,
                    "file_name":      data.get("file_name", ""),
                    "invoice_date":   data.get("invoice_date"),
                    "vendor_name":    data.get("vendor_name"),
                    "invoice_number": data.get("invoice_number"),
                    "currency":       data.get("currency", "ILS"),
                    "total_amount":   data.get("total_amount"),
                    "original_amount":data.get("original_amount"),
                    "vat_amount":     data.get("vat_amount"),
                    "category":       data.get("category", "other"),
                    "ocr_method":     data.get("ocr_method", ""),
                })

            if data.get("ocr_method") == "failed" or not data.get("destination_path"):
                continue

            dest   = Path(data["destination_path"])
            folder = dest.parent
            year   = folder.parent.name
            fname  = folder.name
            try:
                month_num  = fname.split("-")[0]
                month_name = fname.split("-", 1)[1]
            except (IndexError, ValueError):
                month_num, month_name = "00", fname

            key = f"{year}-{month_num}"
            if key not in months:
                months[key] = {
                    "year":       year,
                    "month_num":  month_num,
                    "month_name": f"{month_name} {year}",
                    "count":      0,
                    "total":      0.0,
                    "vat":        0.0,
                    "invoices":   [],
                    "categories": {cat: 0.0 for cat in CATEGORIES},
                }
            m = months[key]
            m["count"] += 1
            m["total"] += data.get("total_amount") or 0.0
            m["vat"]   += data.get("vat_amount")   or 0.0
            cat = data.get("category", "other")
            m["categories"].setdefault(cat, 0.0)
            m["categories"][cat] += data.get("total_amount") or 0.0
            total = data.get("total_amount")
            vat   = data.get("vat_amount")
            m["invoices"].append({
                "file_hash":      fhash,
                "file_name":      data.get("file_name", ""),
                "invoice_date":   data.get("invoice_date"),
                "vendor_name":    data.get("vendor_name"),
                "invoice_number": data.get("invoice_number"),
                "currency":       data.get("currency", "ILS"),
                "original_amount":data.get("original_amount"),
                "exchange_rate":  data.get("exchange_rate"),
                "total_amount":   total,
                "vat_amount":     vat,
                "net":            round(total - vat, 2) if total is not None and vat is not None else None,
                "category":       cat,
                "ocr_method":     data.get("ocr_method", ""),
            })

        for m in months.values():
            m["invoices"].sort(key=lambda i: i.get("invoice_date") or "")
            m["total"] = round(m["total"], 2)
            m["vat"]   = round(m["vat"],   2)

        result["entities"][entity_key] = {
            "label":  entity_cfg["label"],
            "months": dict(sorted(months.items())),
            "review": review,
        }
    return result


def _build_review_html(categories: list) -> str:
    """Build the review-panel HTML + JS (plain string, no f-string escaping needed)."""
    cats_json = json.dumps(categories)
    return (
        """
<style>
#review-panel{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);
  z-index:200;overflow-y:auto;padding:32px 16px;}
.rp-inner{max-width:820px;margin:0 auto;background:#fff;border-radius:16px;padding:32px;}
.rp-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;}
.rp-title{font-size:1.15rem;font-weight:700;color:#1F3A5F;}
.rp-close{background:none;border:none;font-size:1.6rem;cursor:pointer;color:#718096;line-height:1;}
.rp-close:hover{color:#2D3748;}
.rc{background:#F7FAFC;border:1px solid #E2E8F0;border-radius:10px;padding:16px 20px;margin-bottom:14px;}
.rc-file{font-weight:700;color:#1F3A5F;font-size:.9rem;margin-bottom:3px;}
.rc-badge{display:inline-block;padding:2px 9px;border-radius:10px;font-size:.72rem;
  font-weight:600;margin-bottom:10px;}
.rc-badge.fail{background:#fde8e8;color:#E74C3C;}
.rc-badge.warn{background:#fef3cd;color:#d97706;}
.rc-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:12px;}
@media(max-width:600px){.rc-grid{grid-template-columns:1fr 1fr;}}
.rc-field label{display:block;font-size:.7rem;color:#718096;font-weight:600;
  margin-bottom:3px;text-transform:uppercase;letter-spacing:.5px;}
.rc-field input,.rc-field select{width:100%;padding:6px 8px;border:1px solid #CBD5E0;
  border-radius:6px;font-size:.85rem;outline:none;background:#fff;}
.rc-field input:focus,.rc-field select:focus{border-color:#1F3A5F;}
.rc-actions{display:flex;gap:8px;justify-content:flex-end;align-items:center;}
.btn-save{background:#1F3A5F;color:#fff;border:none;padding:8px 20px;
  border-radius:8px;cursor:pointer;font-size:.85rem;font-weight:600;}
.btn-save:hover{background:#2a4f7c;}
.btn-skip{background:#fff;color:#718096;border:1px solid #CBD5E0;padding:8px 14px;
  border-radius:8px;cursor:pointer;font-size:.85rem;}
.rc-saved{color:#2ECC71;font-weight:600;font-size:.85rem;}
.rp-empty{text-align:center;padding:48px;color:#A0AEC0;}
.review-btn{background:rgba(255,255,255,.15);border:2px solid rgba(255,255,255,.4);
  color:#fff;padding:7px 16px;border-radius:20px;cursor:pointer;
  font-size:.85rem;font-weight:600;transition:all .2s;margin-left:8px;}
.review-btn:hover{background:rgba(255,255,255,.25);}
.rv-badge{background:#E74C3C;color:#fff;border-radius:10px;padding:1px 7px;
  font-size:.72rem;margin-left:5px;}
</style>

<div id="review-panel">
  <div class="rp-inner">
    <div class="rp-header">
      <div class="rp-title">Review — Incomplete / Failed Invoices</div>
      <button class="rp-close" onclick="closeReview()">&#x2715;</button>
    </div>
    <div id="review-content"></div>
  </div>
</div>

<script>
const CATS = """ + cats_json + """;

function reviewCount(){
  return Object.values(DATA.entities).reduce((n,e)=>n+(e.review?e.review.length:0),0);
}
function updateReviewBadge(){
  var n=reviewCount(), btn=document.getElementById('review-btn');
  if(!btn) return;
  btn.innerHTML='Review'+(n>0?'<span class="rv-badge">'+n+'</span>':'');
}
function openReview(){ renderReview(); document.getElementById('review-panel').style.display='block'; document.body.style.overflow='hidden'; }
function closeReview(){ document.getElementById('review-panel').style.display='none'; document.body.style.overflow=''; }

function renderReview(){
  var items=[];
  for(var ek in DATA.entities){ var rev=DATA.entities[ek].review||[]; rev.forEach(function(i){ items.push(Object.assign({},i,{entity:ek})); }); }
  var el=document.getElementById('review-content');
  if(!items.length){ el.innerHTML='<div class="rp-empty"><h3>&#x2713; All invoices look complete!</h3><p>Nothing needs review.</p></div>'; return; }
  el.innerHTML=items.map(function(inv){
    var h=inv.file_hash;
    var isFail=inv.ocr_method==='failed';
    var catOpts=CATS.map(function(c){ return '<option value="'+c+'"'+(inv.category===c?' selected':'')+'>'+c+'</option>'; }).join('');
    var origVal=inv.original_amount!=null?inv.original_amount:(inv.total_amount||'');
    var curSel=function(c){ return (inv.currency||'ILS')===c?' selected':''; };
    return '<div class="rc" id="rc-'+h+'">'+
      '<div class="rc-file">'+inv.file_name+'</div>'+
      '<span class="rc-badge '+(isFail?'fail':'warn')+'">'+(isFail?'Processing failed':'Incomplete data')+'</span>'+
      '<div class="rc-grid">'+
        '<div class="rc-field" style="grid-column:span 2"><label>Vendor</label><input id="v-'+h+'" type="text" value="'+(inv.vendor_name||'')+'"></div>'+
        '<div class="rc-field"><label>Date</label><input id="d-'+h+'" type="date" value="'+(inv.invoice_date||'')+'"></div>'+
        '<div class="rc-field"><label>Invoice #</label><input id="n-'+h+'" type="text" value="'+(inv.invoice_number||'')+'"></div>'+
        '<div class="rc-field"><label>Currency</label><select id="cur-'+h+'"><option value="ILS"'+curSel('ILS')+'>&#x20AA; ILS</option><option value="USD"'+curSel('USD')+'>$ USD</option><option value="EUR"'+curSel('EUR')+'>&#x20ac; EUR</option><option value="GBP"'+curSel('GBP')+'>&#xa3; GBP</option></select></div>'+
        '<div class="rc-field"><label>Total (in currency)</label><input id="t-'+h+'" type="number" step="0.01" value="'+origVal+'"></div>'+
        '<div class="rc-field"><label>VAT (0 if none)</label><input id="vt-'+h+'" type="number" step="0.01" value="'+(inv.vat_amount!=null?inv.vat_amount:'')+'"></div>'+
        '<div class="rc-field" style="grid-column:span 3"><label>Category</label><select id="cat-'+h+'">'+catOpts+'</select></div>'+
      '</div>'+
      '<div class="rc-actions">'+
        '<span id="status-'+h+'" class="rc-saved" style="display:none">&#x2713; Saved</span>'+
        '<button class="btn-skip" onclick="skipInvoice(\''+inv.entity+'\',\''+h+'\')">Mark OK as-is</button>'+
        '<button class="btn-save" onclick="saveInvoice(\''+inv.entity+'\',\''+h+'\')">Save &amp; Convert</button>'+
      '</div>'+
    '</div>';
  }).join('');
}

function saveInvoice(entity, h){
  var tVal=parseFloat(document.getElementById('t-'+h).value);
  var vtStr=document.getElementById('vt-'+h).value;
  fetch('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
    entity:entity, hash:h,
    vendor_name:   document.getElementById('v-'+h).value   ||null,
    invoice_date:  document.getElementById('d-'+h).value   ||null,
    invoice_number:document.getElementById('n-'+h).value   ||null,
    currency:      document.getElementById('cur-'+h).value,
    total_amount:  isNaN(tVal)?null:tVal,
    vat_amount:    vtStr!==''?parseFloat(vtStr):null,
    category:      document.getElementById('cat-'+h).value,
  })}).then(function(r){return r.json();}).then(function(res){
    if(res.ok){ document.getElementById('status-'+h).style.display='inline'; document.getElementById('rc-'+h).style.opacity='0.5'; setTimeout(function(){location.reload();},1200); }
    else{ alert('Error: '+(res.error||'unknown')); }
  }).catch(function(e){alert('Server error: '+e);});
}
function skipInvoice(entity,h){
  fetch('/api/update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({entity:entity,hash:h,ocr_method:'manual-skip'})}).then(function(){location.reload();});
}
updateReviewBadge();
</script>
"""
    )


def _open_in_browser(path: Path) -> None:
    """Open a local file in the Windows default browser from WSL."""
    try:
        result = subprocess.run(
            ["wslpath", "-w", str(path)],
            capture_output=True, text=True, timeout=5,
        )
        win_path = result.stdout.strip()
        if win_path:
            subprocess.Popen(
                ["powershell.exe", "-NoProfile", "-Command", f"Start-Process '{win_path}'"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
            return
    except Exception:
        pass
    try:
        webbrowser.open(path.as_uri())
    except Exception as e:
        logger.warning(f"Could not auto-open browser: {e}")


def generate_html_dashboard(all_registries: dict, server_mode: bool = False) -> None:
    """Build processed/dashboard.html and open it in the browser."""
    data          = _build_dashboard_data(all_registries)
    data_json     = json.dumps(data,         ensure_ascii=False, indent=2)
    colors_json   = json.dumps(_CAT_COLORS,  ensure_ascii=False)
    generated     = data["generated"]
    review_html   = _build_review_html(CATEGORIES) if server_mode else ""
    review_btn    = '<button class="review-btn" id="review-btn" onclick="openReview()">Review</button>' if server_mode else ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Invoice Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        background: #F0F4F8; color: #2D3748; min-height: 100vh; }}

/* ── Header ── */
header {{ background: #1F3A5F; color: #fff; padding: 0 32px;
          display: flex; align-items: center; justify-content: space-between;
          height: 64px; box-shadow: 0 2px 8px rgba(0,0,0,.35);
          position: sticky; top: 0; z-index: 100; }}
.brand {{ font-size: 1.25rem; font-weight: 700; letter-spacing: .4px; }}
.brand span {{ color: #7EC8E3; }}
.entity-toggle {{ display: flex; gap: 8px; }}
.entity-btn {{ background: transparent; border: 2px solid rgba(255,255,255,.35);
               color: rgba(255,255,255,.75); padding: 7px 20px; border-radius: 20px;
               cursor: pointer; font-size: .9rem; font-weight: 600; transition: all .2s; }}
.entity-btn:hover  {{ border-color: #fff; color: #fff; }}
.entity-btn.active {{ background: #fff; border-color: #fff; color: #1F3A5F; }}

/* ── Layout ── */
.container {{ max-width: 1200px; margin: 0 auto; padding: 32px 24px; }}

/* ── Summary cards ── */
.cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 32px; }}
@media (max-width: 768px) {{ .cards {{ grid-template-columns: repeat(2, 1fr); }} }}
.card {{ background: #fff; border-radius: 12px; padding: 20px 24px;
         box-shadow: 0 1px 4px rgba(0,0,0,.08); border-top: 4px solid; }}
.card:nth-child(1) {{ border-color: #1F3A5F; }}
.card:nth-child(2) {{ border-color: #2ECC71; }}
.card:nth-child(3) {{ border-color: #E74C3C; }}
.card:nth-child(4) {{ border-color: #3498DB; }}
.card-label {{ font-size: .75rem; text-transform: uppercase; letter-spacing: .8px;
               color: #718096; font-weight: 600; margin-bottom: 8px; }}
.card-value {{ font-size: 1.7rem; font-weight: 700; color: #1A202C; }}
.card-sub   {{ font-size: .8rem; color: #A0AEC0; margin-top: 4px; }}

/* ── Section title ── */
.section-title {{ font-size: .8rem; font-weight: 700; color: #718096;
                  text-transform: uppercase; letter-spacing: .8px; margin-bottom: 10px; }}

/* ── Year tabs ── */
.year-tabs {{ display: flex; gap: 8px; margin-bottom: 28px; flex-wrap: wrap; }}
.year-btn {{ background: #fff; border: 2px solid #CBD5E0; color: #4A5568;
             padding: 8px 22px; border-radius: 8px; cursor: pointer;
             font-size: .95rem; font-weight: 600; transition: all .2s; }}
.year-btn:hover  {{ border-color: #1F3A5F; color: #1F3A5F; }}
.year-btn.active {{ background: #1F3A5F; border-color: #1F3A5F; color: #fff; }}

/* ── Month grid ── */
.month-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }}
@media (max-width: 800px) {{ .month-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
.month-card {{ background: #fff; border-radius: 12px; padding: 20px;
               cursor: pointer; box-shadow: 0 1px 4px rgba(0,0,0,.08);
               border: 2px solid transparent; transition: all .2s; }}
.month-card:hover {{ border-color: #1F3A5F; transform: translateY(-2px);
                     box-shadow: 0 6px 16px rgba(31,58,95,.15); }}
.month-card.empty {{ opacity: .4; cursor: default; pointer-events: none; }}
.mc-name  {{ font-weight: 700; font-size: 1rem; color: #2D3748; margin-bottom: 10px; }}
.mc-count {{ font-size: .8rem; color: #718096; }}
.mc-total {{ font-size: 1.2rem; font-weight: 700; color: #1F3A5F; margin-top: 6px; }}

/* ── Breadcrumb ── */
.breadcrumb {{ display: flex; align-items: center; gap: 8px; margin-bottom: 20px;
               font-size: .9rem; color: #718096; }}
.bc-link {{ color: #1F3A5F; font-weight: 600; cursor: pointer; text-decoration: none; }}
.bc-link:hover {{ text-decoration: underline; }}
.bc-sep {{ color: #CBD5E0; }}

/* ── Month detail ── */
.month-detail {{ display: grid; grid-template-columns: 1fr 300px; gap: 24px; align-items: start; }}
@media (max-width: 900px) {{ .month-detail {{ grid-template-columns: 1fr; }} }}

/* ── Invoice table ── */
.table-wrap {{ background: #fff; border-radius: 12px; overflow: hidden;
               box-shadow: 0 1px 4px rgba(0,0,0,.08); overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; font-size: .88rem; }}
th {{ background: #1F3A5F; color: #fff; padding: 12px 14px;
      text-align: left; font-weight: 600; white-space: nowrap; }}
th.r {{ text-align: right; }}
td {{ padding: 11px 14px; border-bottom: 1px solid #EDF2F7; }}
td.r {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
tr:nth-child(even) td {{ background: #F7FAFC; }}
tr:last-child td {{ border-bottom: none; }}
.tfoot td {{ background: #1F3A5F !important; color: #fff; font-weight: 700; }}
.badge {{ display: inline-block; padding: 2px 10px; border-radius: 12px;
          font-size: .73rem; font-weight: 600; }}

/* ── Chart card ── */
.chart-card {{ background: #fff; border-radius: 12px; padding: 24px;
               box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
.chart-title {{ font-size: .95rem; font-weight: 700; color: #2D3748;
                margin-bottom: 16px; text-align: center; }}
.legend-item {{ display: flex; align-items: center; gap: 8px; padding: 5px 0;
                font-size: .82rem; color: #4A5568; }}
.legend-dot {{ width: 11px; height: 11px; border-radius: 50%; flex-shrink: 0; }}
.legend-val {{ margin-left: auto; font-weight: 600; color: #2D3748; }}

/* ── Empty / misc ── */
.empty-state {{ text-align: center; padding: 60px 20px; color: #A0AEC0; }}
.no-data {{ color: #CBD5E0; font-style: italic; }}
footer {{ text-align: center; padding: 24px; color: #A0AEC0;
          font-size: .78rem; margin-top: 48px; }}
</style>
</head>
<body>

<header>
  <div class="brand">Invoice <span>Dashboard</span></div>
  <div style="display:flex;gap:8px;align-items:center">
    <div class="entity-toggle" id="entity-toggle"></div>
    {review_btn}
  </div>
</header>

<div class="container">
  <div class="cards"        id="summary-cards"></div>
  <div class="section-title">Year</div>
  <div class="year-tabs"    id="year-tabs"></div>
  <div                      id="main-content"></div>
</div>

<footer>Generated {generated} &nbsp;·&nbsp; Invoice Processor</footer>

<script>
const DATA       = {data_json};
const CAT_COLORS = {colors_json};
const MONTH_NAMES = ['','January','February','March','April','May','June',
                     'July','August','September','October','November','December'];

let state = {{ entity: Object.keys(DATA.entities)[0], year: null, month: null }};
let pieChart = null;

// ── Helpers ────────────────────────────────────────────────────────────────

function fmt(n) {{
  if (n == null) return '<span class="no-data">—</span>';
  return '₪\u202F' + n.toLocaleString('he-IL', {{minimumFractionDigits:2, maximumFractionDigits:2}});
}}
function fmtPlain(n) {{
  if (n == null) return '—';
  return '₪\u202F' + n.toLocaleString('he-IL', {{minimumFractionDigits:2, maximumFractionDigits:2}});
}}
function ed()   {{ return DATA.entities[state.entity] || {{label:'', months:{{}}}}; }}
function years()  {{ return [...new Set(Object.values(ed().months).map(m => m.year))].sort(); }}
function monthsForYear(y) {{
  return Object.values(ed().months).filter(m => m.year === y)
               .sort((a,b) => a.month_num.localeCompare(b.month_num));
}}
function summary(year) {{
  const ms = year ? monthsForYear(year) : Object.values(ed().months);
  return ms.reduce((a,m) => ({{ count: a.count+m.count, total: a.total+m.total, vat: a.vat+m.vat }}),
                   {{ count:0, total:0, vat:0 }});
}}

// ── Render: entity toggle ──────────────────────────────────────────────────

function renderEntityToggle() {{
  document.getElementById('entity-toggle').innerHTML =
    Object.entries(DATA.entities).map(([k,e]) =>
      `<button class="entity-btn ${{state.entity===k?'active':''}}"
               onclick="setEntity('${{k}}')">${{e.label}}</button>`
    ).join('');
}}

// ── Render: summary cards ──────────────────────────────────────────────────

function renderCards() {{
  const s   = summary(state.year);
  const net = s.total - s.vat;
  const scope = state.year || 'All time';
  document.getElementById('summary-cards').innerHTML = `
    <div class="card">
      <div class="card-label">Total Invoices</div>
      <div class="card-value">${{s.count}}</div>
      <div class="card-sub">${{scope}}</div>
    </div>
    <div class="card">
      <div class="card-label">Total Spend</div>
      <div class="card-value">${{fmtPlain(s.total)}}</div>
      <div class="card-sub">incl. VAT</div>
    </div>
    <div class="card">
      <div class="card-label">Total VAT (מע"מ)</div>
      <div class="card-value">${{fmtPlain(s.vat)}}</div>
      <div class="card-sub">&nbsp;</div>
    </div>
    <div class="card">
      <div class="card-label">Net Amount</div>
      <div class="card-value">${{fmtPlain(net)}}</div>
      <div class="card-sub">excl. VAT</div>
    </div>`;
}}

// ── Render: year tabs ──────────────────────────────────────────────────────

function renderYearTabs() {{
  const ys = years();
  if (!state.year && ys.length) state.year = ys[ys.length - 1];
  document.getElementById('year-tabs').innerHTML = ys.length
    ? ys.map(y => `<button class="year-btn ${{state.year===y?'active':''}}"
                           onclick="setYear('${{y}}')">${{y}}</button>`).join('')
    : '<span class="no-data">No data yet</span>';
}}

// ── Render: month grid ─────────────────────────────────────────────────────

function renderMonthGrid() {{
  const map = {{}};
  if (state.year) monthsForYear(state.year).forEach(m => {{ map[parseInt(m.month_num)] = m; }});
  const cards = Array.from({{length:12}}, (_,i) => {{
    const m = map[i+1], name = MONTH_NAMES[i+1];
    if (!m) return `<div class="month-card empty">
                      <div class="mc-name">${{name}}</div>
                      <div class="mc-count">No invoices</div>
                      <div class="mc-total">—</div>
                    </div>`;
    return `<div class="month-card" onclick="setMonth('${{m.month_num}}')">
              <div class="mc-name">${{name}}</div>
              <div class="mc-count">${{m.count}} invoice${{m.count!==1?'s':''}}</div>
              <div class="mc-total">${{fmtPlain(m.total)}}</div>
            </div>`;
  }}).join('');
  document.getElementById('main-content').innerHTML = `<div class="month-grid">${{cards}}</div>`;
}}

// ── Render: month detail ───────────────────────────────────────────────────

function renderMonthDetail() {{
  const key = `${{state.year}}-${{state.month}}`;
  const m   = ed().months[key];
  if (!m) {{ renderMonthGrid(); return; }}

  const rows = m.invoices.map(inv => `
    <tr>
      <td>${{inv.invoice_date || '<span class="no-data">—</span>'}}</td>
      <td>${{inv.vendor_name  || '<span class="no-data">—</span>'}}</td>
      <td>${{inv.invoice_number || '<span class="no-data">—</span>'}}</td>
      <td class="r">${{fmt(inv.total_amount)}}</td>
      <td class="r">${{fmt(inv.vat_amount)}}</td>
      <td class="r">${{fmt(inv.net)}}</td>
      <td><span class="badge"
            style="background:${{(CAT_COLORS[inv.category]||'#95A5A6')}}22;
                   color:${{CAT_COLORS[inv.category]||'#95A5A6'}}"
          >${{inv.category}}</span></td>
    </tr>`).join('');

  const net = m.total - m.vat;
  const activeCats = Object.entries(m.categories)
                           .filter(([,v]) => v > 0)
                           .sort((a,b) => b[1]-a[1]);
  const legend = activeCats.map(([cat,val]) => `
    <div class="legend-item">
      <div class="legend-dot" style="background:${{CAT_COLORS[cat]||'#95A5A6'}}"></div>
      <span>${{cat}}</span>
      <span class="legend-val">${{fmtPlain(val)}}</span>
    </div>`).join('');

  document.getElementById('main-content').innerHTML = `
    <div class="breadcrumb">
      <span class="bc-link" onclick="setMonth(null)">${{state.year}}</span>
      <span class="bc-sep">›</span>
      <span>${{m.month_name}}</span>
    </div>
    <div class="month-detail">
      <div class="table-wrap">
        <table>
          <thead><tr>
            <th>Date</th><th>Vendor</th><th>Invoice #</th>
            <th class="r">Total (₪)</th><th class="r">VAT (₪)</th>
            <th class="r">Net (₪)</th><th>Category</th>
          </tr></thead>
          <tbody>${{rows}}</tbody>
          <tfoot><tr class="tfoot">
            <td colspan="3">TOTAL (${{m.count}} invoices)</td>
            <td class="r">${{fmtPlain(m.total)}}</td>
            <td class="r">${{fmtPlain(m.vat)}}</td>
            <td class="r">${{fmtPlain(net)}}</td>
            <td></td>
          </tr></tfoot>
        </table>
      </div>
      <div class="chart-card">
        <div class="chart-title">Spending by Category</div>
        <canvas id="pieCanvas" width="252" height="252"></canvas>
        <div style="margin-top:16px">${{legend}}</div>
      </div>
    </div>`;

  if (pieChart) {{ pieChart.destroy(); pieChart = null; }}
  if (activeCats.length) {{
    pieChart = new Chart(document.getElementById('pieCanvas').getContext('2d'), {{
      type: 'doughnut',
      data: {{
        labels:   activeCats.map(([c]) => c),
        datasets: [{{ data: activeCats.map(([,v]) => v),
                      backgroundColor: activeCats.map(([c]) => CAT_COLORS[c]||'#95A5A6'),
                      borderWidth: 2, borderColor: '#fff' }}]
      }},
      options: {{
        responsive: false,
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{ callbacks: {{ label: ctx =>
            ` ${{ctx.label}}: ₪${{ctx.parsed.toLocaleString('he-IL',{{minimumFractionDigits:2}})}}`
          }} }}
        }}
      }}
    }});
  }}
}}

// ── State setters ──────────────────────────────────────────────────────────

function setEntity(key) {{
  const ys = years();
  state = {{ entity: key, year: null, month: null }};
  const nys = years();
  if (nys.length) state.year = nys[nys.length-1];
  render();
}}
function setYear(y)  {{ state.year = y; state.month = null; render(); }}
function setMonth(m) {{ state.month = m; renderCards(); renderMain(); }}

// ── Main render ────────────────────────────────────────────────────────────

function renderMain() {{
  if (!state.year) {{
    document.getElementById('main-content').innerHTML =
      '<div class="empty-state"><h3>No invoices yet</h3>' +
      '<p>Drop files into incoming_invoices/ohad or incoming_invoices/cril-tech and run the processor.</p></div>';
    return;
  }}
  state.month ? renderMonthDetail() : renderMonthGrid();
}}

function render() {{
  renderEntityToggle();
  renderCards();
  renderYearTabs();
  renderMain();
}}

render();
</script>
{review_html}
</body>
</html>"""

    DASHBOARD_PATH.parent.mkdir(parents=True, exist_ok=True)
    DASHBOARD_PATH.write_text(html, encoding="utf-8")
    logger.info(f"HTML dashboard → {DASHBOARD_PATH.relative_to(BASE_DIR)}")
    if not server_mode:
        _open_in_browser(DASHBOARD_PATH)


# ── Local review server ───────────────────────────────────────────────────────

def run_server(port: int = 8080) -> None:
    """Start a local HTTP server for the interactive review dashboard."""
    import http.server

    class ReviewHandler(http.server.BaseHTTPRequestHandler):
        def log_message(self, format, *args): pass  # suppress default access logs

        def do_GET(self):
            if self.path in ("/", "/dashboard"):
                all_reg = {
                    key: (cfg, load_registry(cfg["registry_path"]))
                    for key, cfg in ENTITIES.items()
                }
                generate_html_dashboard(all_reg, server_mode=True)
                content = DASHBOARD_PATH.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", len(content))
                self.end_headers()
                self.wfile.write(content)
            else:
                self.send_response(404)
                self.end_headers()

        def do_POST(self):
            if self.path != "/api/update":
                self.send_response(404); self.end_headers(); return
            length = int(self.headers.get("Content-Length", 0))
            body   = self.rfile.read(length)
            try:
                update   = json.loads(body)
                entity   = update.pop("entity", "ohad")
                fhash    = update.pop("hash")
                cfg      = ENTITIES[entity]
                registry = load_registry(cfg["registry_path"])

                if fhash not in registry:
                    self._json(404, {"ok": False, "error": "hash not found"}); return

                entry    = registry[fhash]
                currency = (update.get("currency") or entry.get("currency") or "ILS").upper()
                total    = update.get("total_amount")
                vat      = update.get("vat_amount")

                # Convert foreign currency → ILS when user enters the amount
                if currency != "ILS" and total is not None:
                    rate = get_boi_rate(currency, update.get("invoice_date") or entry.get("invoice_date"))
                    if rate:
                        update["original_amount"] = total
                        update["exchange_rate"]   = rate
                        update["total_amount"]    = round(total * rate, 2)
                        if vat is not None:
                            update["vat_amount"]  = round(vat * rate, 2)
                        logger.info(f"Server: {currency}→ILS @ {rate}: {total} → ₪{update['total_amount']}")

                # VAT fallback for ILS invoices
                final_total = update.get("total_amount", entry.get("total_amount"))
                final_vat   = update.get("vat_amount",   entry.get("vat_amount"))
                if final_vat is None and final_total is not None and currency == "ILS":
                    update["vat_amount"] = round(final_total * 18 / 118, 2)

                # Apply update
                entry.update(update)
                entry["currency"] = currency

                # Move file if invoice_date changed to a different month
                old_dest = Path(entry.get("destination_path", ""))
                new_date = entry.get("invoice_date")
                if old_dest.exists() and new_date:
                    new_folder = month_folder_for(new_date, cfg["processed_base"])
                    if new_folder != old_dest.parent:
                        new_folder.mkdir(parents=True, exist_ok=True)
                        new_dest = new_folder / old_dest.name
                        ctr = 1
                        while new_dest.exists():
                            new_dest = new_folder / f"{old_dest.stem}_{ctr}{old_dest.suffix}"; ctr += 1
                        shutil.move(str(old_dest), new_dest)
                        entry["destination_path"] = str(new_dest)
                        logger.info(f"Moved {old_dest.name} → {new_dest.relative_to(BASE_DIR)}")
                        old_month_invoices = [
                            InvoiceData(**d) for d in registry.values()
                            if d.get("destination_path") and
                               Path(d["destination_path"]).parent == old_dest.parent
                        ]
                        if old_dest.parent.exists():
                            generate_excel_report(old_dest.parent, old_month_invoices)

                registry[fhash] = entry
                save_registry(registry, cfg["registry_path"])

                # Regenerate month report
                dest_path = Path(entry.get("destination_path", ""))
                if dest_path.exists():
                    mf = dest_path.parent
                    all_month = [
                        InvoiceData(**d) for d in registry.values()
                        if d.get("destination_path") and Path(d["destination_path"]).parent == mf
                    ]
                    generate_excel_report(mf, all_month)

                # Regenerate dashboard
                all_reg = {
                    key: (cfg2, load_registry(cfg2["registry_path"]))
                    for key, cfg2 in ENTITIES.items()
                }
                generate_html_dashboard(all_reg, server_mode=True)
                self._json(200, {"ok": True})

            except Exception as e:
                logger.exception("Server update error")
                self._json(500, {"ok": False, "error": str(e)})

        def _json(self, code, obj):
            body = json.dumps(obj).encode()
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", len(body))
            self.end_headers()
            self.wfile.write(body)

    url = f"http://localhost:{port}"
    try:
        subprocess.Popen(
            ["powershell.exe", "-NoProfile", "-Command", f"Start-Process '{url}'"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
    except Exception:
        pass
    server = http.server.ThreadingHTTPServer(("localhost", port), ReviewHandler)
    logger.info(f"Review server → {url}  (Ctrl+C to stop)")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        logger.info("Server stopped.")


# ── Core processor ────────────────────────────────────────────────────────────

def process_file(
    file_path:      Path,
    client:         anthropic.Anthropic,
    registry:       dict,
    registry_path:  Path,
    processed_base: Path,
) -> Optional[InvoiceData]:
    fhash    = file_sha256(file_path)
    existing = registry.get(fhash)

    if existing and existing.get("ocr_method") != "failed":
        logger.info(f"Skip (already processed): {file_path.name}")
        return None

    existing_dest: Optional[Path] = None
    if existing:
        prev = existing.get("destination_path", "")
        if prev and Path(prev).exists():
            existing_dest = Path(prev)

    logger.info(f"Processing: {file_path.name}")
    extracted: dict = {}
    method = "unknown"

    try:
        import pytesseract  # noqa
        ocr_text, confidence = tesseract_ocr(file_path)
        if validate_ocr(ocr_text, confidence):
            logger.info(f"Tesseract OK ({confidence:.0f}% confidence)")
            extracted = extract_from_text(ocr_text)
            method    = "tesseract"
        else:
            logger.info(f"Tesseract validation failed ({confidence:.0f}%), falling back to Claude API")
    except ImportError:
        logger.info("pytesseract not installed – using Claude API")
    except Exception as e:
        logger.warning(f"Tesseract error: {e} – falling back to Claude API")

    if method == "unknown":
        try:
            extracted = extract_with_claude(file_path, client)
            method    = "claude"
        except Exception as e:
            logger.error(f"Claude API failed for {file_path.name}: {e}")
            extracted = {}
            method    = "failed"

    total    = extracted.get("total_amount")
    vat      = extracted.get("vat_amount")
    currency = (extracted.get("currency") or "ILS").upper()
    original_amount: Optional[float] = None
    exchange_rate:   Optional[float] = None

    # Convert foreign currency → ILS using Bank of Israel שער יציג
    if currency != "ILS" and total is not None:
        rate = get_boi_rate(currency, extracted.get("invoice_date"))
        if rate:
            original_amount = total
            total           = round(total * rate, 2)
            vat             = round(vat   * rate, 2) if vat is not None else None
            exchange_rate   = rate
            logger.info(f"  {currency}→ILS @ {rate}: original={original_amount}, total=₪{total}")
        else:
            logger.warning(f"  Could not convert {currency}→ILS, keeping raw value")

    # VAT fallback
    if vat is None and total is not None:
        if currency == "ILS":
            # Israeli invoice with no explicit VAT → assume 18% included
            vat = round(total * 18 / 118, 2)
            logger.info(f"  VAT not found – calculated from total at 18%: ₪{vat}")
        else:
            # Foreign invoice with unknown VAT → 0 (no Israeli VAT)
            vat = 0.0

    invoice = InvoiceData(
        file_name       = file_path.name,
        file_hash       = fhash,
        invoice_date    = extracted.get("invoice_date"),
        vendor_name     = extracted.get("vendor_name"),
        invoice_number  = extracted.get("invoice_number"),
        total_amount    = total,
        vat_amount      = vat,
        currency        = currency,
        original_amount = original_amount,
        exchange_rate   = exchange_rate,
        category        = extracted.get("category") or "other",
        ocr_method      = method,
    )

    if existing_dest:
        dest = existing_dest
        logger.info(f"  reusing existing copy: {dest.relative_to(BASE_DIR)}")
    else:
        dest = copy_to_processed(file_path, invoice, processed_base)
    invoice.destination_path = str(dest)

    registry[fhash] = asdict(invoice)
    save_registry(registry, registry_path)
    return invoice


# ── Main entry point ──────────────────────────────────────────────────────────

def run(regenerate_reports: bool = False) -> None:
    # Ensure all directories exist
    for cfg in ENTITIES.values():
        cfg["incoming_dir"].mkdir(parents=True, exist_ok=True)
        cfg["processed_base"].mkdir(parents=True, exist_ok=True)

    # Load all registries once
    all_registries: dict = {
        key: (cfg, load_registry(cfg["registry_path"]))
        for key, cfg in ENTITIES.items()
    }

    if regenerate_reports:
        logger.info("Regenerating all monthly Excel reports …")
        for entity_key, (cfg, registry) in all_registries.items():
            month_groups: dict[str, list[InvoiceData]] = {}
            for data in registry.values():
                dest = data.get("destination_path", "")
                if not dest:
                    continue
                folder = str(Path(dest).parent)
                month_groups.setdefault(folder, []).append(InvoiceData(**data))
            for folder_str, invoices in month_groups.items():
                folder = Path(folder_str)
                if folder.exists():
                    generate_excel_report(folder, invoices)
        generate_html_dashboard(all_registries)
        logger.info("Done.")
        return

    client  = anthropic.Anthropic()
    any_new = False

    for entity_key, cfg in ENTITIES.items():
        _, registry = all_registries[entity_key]

        files = sorted(
            f for f in cfg["incoming_dir"].iterdir()
            if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
        )
        if not files:
            logger.info(f"[{entity_key}] No files in {cfg['incoming_dir'].relative_to(BASE_DIR)}")
            continue

        logger.info(f"[{entity_key}] Found {len(files)} file(s)")
        new_invoices: list[InvoiceData] = []

        for file_path in files:
            inv = process_file(
                file_path, client, registry,
                cfg["registry_path"], cfg["processed_base"],
            )
            if inv:
                new_invoices.append(inv)

        if not new_invoices:
            logger.info(f"[{entity_key}] No new invoices.")
            continue

        any_new = True

        affected: set[Path] = {Path(inv.destination_path).parent for inv in new_invoices}
        for month_folder in affected:
            all_month = [
                InvoiceData(**data)
                for data in registry.values()
                if data.get("destination_path")
                and Path(data["destination_path"]).parent == month_folder
            ]
            generate_excel_report(month_folder, all_month)

        sep = "─" * 62
        print(f"\n{sep}")
        print(f"  [{cfg['label']}]  Processed {len(new_invoices)} invoice(s)")
        print(sep)
        for inv in new_invoices:
            total_str = f"₪{inv.total_amount:,.2f}" if inv.total_amount else "₪?"
            vat_str   = f"₪{inv.vat_amount:,.2f}"   if inv.vat_amount   else "₪?"
            print(f"  {inv.file_name}")
            print(f"    Date: {inv.invoice_date or 'N/A':12}  Vendor: {inv.vendor_name or 'N/A'}")
            print(f"    Total: {total_str:12}  VAT: {vat_str:10}  Cat: {inv.category}")
            print(f"    Method: {inv.ocr_method}")
        print(sep + "\n")

    if any_new:
        # Reload updated registries before building dashboard
        all_registries = {
            key: (cfg, load_registry(cfg["registry_path"]))
            for key, cfg in ENTITIES.items()
        }
        generate_html_dashboard(all_registries)


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Invoice Processor – Ohad Kazoom (עוסק מורשה) & Cril-Tech"
    )
    parser.add_argument(
        "--regenerate-reports", action="store_true",
        help="Rebuild all Excel reports + HTML dashboard without re-processing files",
    )
    parser.add_argument(
        "--serve", action="store_true",
        help="Start local review server at http://localhost:8080",
    )
    parser.add_argument(
        "--port", type=int, default=8080,
        help="Port for --serve mode (default: 8080)",
    )
    args = parser.parse_args()
    if args.serve:
        run_server(port=args.port)
    else:
        run(regenerate_reports=args.regenerate_reports)
